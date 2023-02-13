from django.core.management.base import BaseCommand, CommandError
from app_webshop.settings import DATA_DIR
from market.models import Category, Product
from openpyxl import load_workbook

class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        print('[+] Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        category = None

        print('[+] Start importing from excel... ' + str(DATA_DIR))
        wb = load_workbook(DATA_DIR + '/price.xlsx')
        worksheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        for idx in range(4, worksheet.max_row + 1):
            id = worksheet.cell(row=idx, column=2).value
            name = worksheet.cell(row=idx, column=3).value
            units = worksheet.cell(row=idx, column=4).value
            balance = worksheet.cell(row=idx, column=5).value
            price = worksheet.cell(row=idx, column=6).value
            print('Название: ' + str(name) + ' ' 'ед.изм.: ' +str(units) + ' ' + 'кол-во: '
                + str(balance) + ' ' + 'Цена за ед.: ' + str(price))
            if id is None:
                print('[+] Create a new category')
                category = Category()
                category.name = name
                category.save()
            else:
                print('[+] Create a new good')
                if category:
                    product = Product()
                    product.name = name
                    product.category = category
                    product.save()



